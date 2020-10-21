# -*- coding: utf-8 -*-
"""
Created on Fri Jan  3 15:56:47 2020

@author: hanbing.guo
"""

import openpyxl as xl
import numpy as np
from datetime import timedelta
from datetime import date, datetime
import os
import shutil
from collections import namedtuple


TimePoint = namedtuple('TimePoint', ['h', 'm'])
late_point = TimePoint(8,34)
early_point = TimePoint(17,45)
over_point = TimePoint(20,30)

wb = xl.load_workbook('new.xlsx')

sheet = []
useful_idx = [0,2,3,4, 5,6]
for st in wb:
    print(st.title)
    print(st.max_row, ', ', st.max_column)
    for row in st.rows:
        row_list = [row[e].value for e in useful_idx]
        sheet.append(row_list)
        
del sheet[0]

#d1 = datetime.strptime(sheet[0][2], "%Y-%m-%d %H:%M")
#d2 = datetime.strptime(sheet[1][2], "%Y-%m-%d %H:%M")
#d3 = datetime.strptime(sheet[7][2], "%Y-%m-%d %H:%M")
#print(d2-d1)
#print(d1-d2)
#print(d1 > d2)
#print(d1.day == d3.day)
for row in sheet:
    row[2] = datetime.strptime(row[2], "%Y-%m-%d %H:%M")
    
late = []
early = []
overtime = []

current_id = -1
current_name = ''
current_time = None
min_time = None
max_time = None
day_check = []

for i, row in enumerate(sheet):
    num, name, check_time = row[0:3]
    if num == current_id:
        if current_time.day == check_time.day:
            if min_time and max_time:
                if check_time < min_time:
                    min_time = check_time
                if max_time < check_time:
                    max_time = check_time
            else:
                min_time = check_time
                max_time = check_time
        else:
            day_check.append([i + 1, current_id, current_name, min_time, max_time])
            min_time = check_time
            max_time = check_time
            
        current_time = check_time
    else:
        if int(current_id) >= 0:
            day_check.append([i + 1, current_id, current_name, min_time, max_time])
            min_time = check_time
            max_time = check_time
            
        current_id = num
        current_name = name
        current_time = check_time
        min_time = check_time
        max_time = check_time
        
for row in day_check:
    in_time, out_time = row[3:5]
    if in_time.hour >= late_point.h and in_time.minute > late_point.m:
        late.append(row)
    if out_time.hour <= early_point.h and out_time.minute < early_point.m:
        early.append(row)
    if out_time.hour >= over_point.h and out_time.minute >= over_point.m:
        overtime.append(row)
    if in_time.isoweekday() > 5 and (out_time - in_time).seconds > 21600:
        overtime.append(row)

    
print('Late: ', len(late))
for row in late:
    print(row)
    
print('Early: ', len(early))
for row in early:
    print(row)
    
print('Overtime: ', len(overtime))
for row in overtime:
    print(row)


with open('check.csv', 'w') as fh:
    fh.writelines('line,ID,name,day,in,out\n')
    fh.writelines('late,' +str(len(late))+'\n')
    for row in late:
        in_time, out_time = row[3:5]
        line = ','.join([str(row[0]), str(row[1]), row[2], 
                        '-'.join([str(in_time.year), str(in_time.month), str(in_time.day)])
                        , ':'.join([str(in_time.hour), str(in_time.minute)])
                        , ':'.join([str(out_time.hour), str(out_time.minute)])])
        line = line + '\n'
        fh.writelines(line)
 
    fh.writelines('early,' +str(len(early))+'\n')
    for row in early:
        in_time, out_time = row[3:5]
        line = ','.join([str(row[0]), str(row[1]), row[2], 
                        '-'.join([str(in_time.year), str(in_time.month), str(in_time.day)])
                        , ':'.join([str(in_time.hour), str(in_time.minute)])
                        , ':'.join([str(out_time.hour), str(out_time.minute)])])
        line = line + '\n'
        fh.writelines(line)
        
    fh.writelines('overtime,' +str(len(overtime))+'\n')
    for row in overtime:
        in_time, out_time = row[3:5]
        line = ','.join([str(row[0]), str(row[1]), row[2], 
                        '-'.join([str(in_time.year), str(in_time.month), str(in_time.day)])
                        , ':'.join([str(in_time.hour), str(in_time.minute)])
                        , ':'.join([str(out_time.hour), str(out_time.minute)])])
        line = line + '\n'
        fh.writelines(line)